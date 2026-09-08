import csv
from dataclasses import dataclass
from datetime import date, datetime
from html import escape
from io import BytesIO, StringIO
from typing import Literal, Sequence

from fastapi import HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


ReportFormat = Literal["csv", "xlsx", "pdf"]


@dataclass(frozen=True)
class TabularReportSpec:
    title: str
    filename_prefix: str
    sheet_name: str
    headers: tuple[str, ...]
    pdf_columns: tuple[int, ...]
    pdf_widths_mm: tuple[float, ...]
    currency_columns: tuple[int, ...] = ()

    def __post_init__(self) -> None:
        if len(self.pdf_columns) != len(self.pdf_widths_mm):
            raise ValueError("PDF columns and widths must have equal length.")
        if not self.headers:
            raise ValueError("A tabular report requires at least one column.")


def period_label(from_date: date | None, to_date: date | None) -> str:
    if from_date and to_date:
        return f"{from_date:%d %b %Y} - {to_date:%d %b %Y}"
    if from_date:
        return f"From {from_date:%d %b %Y}"
    if to_date:
        return f"Through {to_date:%d %b %Y}"
    return "All available data"


def safe_csv_cell(value: object) -> str:
    """Prevent spreadsheet formula execution in downloaded files."""
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@", "\t", "\r")):
        return f"'{text}"
    return text


def export_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return value
    return safe_csv_cell(value)


def serialize_rows(rows: Sequence[Sequence[object]]) -> list[list[object]]:
    return [[export_cell(value) for value in row] for row in rows]


def _validate_rows(
    rows: Sequence[Sequence[object]],
    spec: TabularReportSpec,
) -> None:
    expected = len(spec.headers)
    if any(len(row) != expected for row in rows):
        raise ValueError(
            f"Every {spec.title} row must contain {expected} columns."
        )


def _build_csv(
    rows: Sequence[Sequence[object]],
    headers: Sequence[str],
) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(
        [safe_csv_cell(value) for value in row] for row in rows
    )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _build_xlsx(
    rows: Sequence[Sequence[object]],
    metadata: Sequence[tuple[str, str]],
    spec: TabularReportSpec,
) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append([spec.title])
    summary["A1"].font = Font(bold=True, size=16)
    for label, value in metadata:
        summary.append([safe_csv_cell(label), safe_csv_cell(value)])
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 48

    detail = workbook.create_sheet(spec.sheet_name[:31])
    detail.append(list(spec.headers))
    for row in rows:
        detail.append([export_cell(value) for value in row])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    header_fill = PatternFill("solid", fgColor="1B5E20")
    for cell in detail[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for column_index, header in enumerate(spec.headers, start=1):
        width = min(max(len(header) + 2, 14), 34)
        detail.column_dimensions[get_column_letter(column_index)].width = width
    for row_index in range(2, detail.max_row + 1):
        for column_index in spec.currency_columns:
            detail.cell(row=row_index, column=column_index).number_format = (
                '₹#,##0.00'
            )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_pdf(
    rows: Sequence[Sequence[object]],
    metadata: Sequence[tuple[str, str]],
    spec: TabularReportSpec,
) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A3),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=spec.title,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(escape(spec.title), styles["Title"]),
        Spacer(1, 4 * mm),
    ]
    story.extend(
        Paragraph(
            f"<b>{escape(str(label))}:</b> {escape(str(value))}",
            styles["BodyText"],
        )
        for label, value in metadata
    )
    story.append(Spacer(1, 5 * mm))

    table_data = [
        [spec.headers[index] for index in spec.pdf_columns],
        *[
            [str(export_cell(row[index])) for index in spec.pdf_columns]
            for row in rows
        ],
    ]
    table = Table(
        table_data,
        colWidths=[width * mm for width in spec.pdf_widths_mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E20")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F8FAFC")],
                ),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)
    document.build(story)
    return output.getvalue()


def build_tabular_report_response(
    rows: list[list[object]],
    *,
    metadata: list[tuple[str, str]],
    export_format: ReportFormat,
    snapshot: datetime,
    spec: TabularReportSpec,
    row_limit: int,
    pdf_row_limit: int,
) -> Response:
    _validate_rows(rows, spec)
    if len(rows) > row_limit:
        raise HTTPException(
            status_code=413,
            detail=(
                f"The export exceeds {row_limit:,} rows. "
                "Use a smaller date range."
            ),
        )
    if export_format == "pdf" and len(rows) > pdf_row_limit:
        raise HTTPException(
            status_code=413,
            detail=(
                f"PDF exports support up to {pdf_row_limit:,} rows. "
                "Use a smaller date range or download XLSX/CSV."
            ),
        )

    if export_format == "xlsx":
        content = _build_xlsx(rows, metadata, spec)
        media_type = (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    elif export_format == "pdf":
        content = _build_pdf(rows, metadata, spec)
        media_type = "application/pdf"
    else:
        content = _build_csv(rows, spec.headers)
        media_type = "text/csv; charset=utf-8"

    filename = (
        f"{spec.filename_prefix}-{snapshot:%Y-%m-%d-%H%M%S}-IST."
        f"{export_format}"
    )
    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Report-Snapshot": snapshot.isoformat(),
            "X-Report-Row-Count": str(len(rows)),
        },
    )
