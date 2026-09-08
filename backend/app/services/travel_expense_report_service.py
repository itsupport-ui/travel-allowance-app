import calendar
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from html import escape
from io import BytesIO, StringIO
from typing import Literal

import csv as csv_module

from fastapi import HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.orm import Session

from app.models.claim import Claim
from app.models.doctor import Doctor
from app.models.doctor_expense import DoctorExpense
from app.models.doctor_visit import DoctorVisit
from app.models.travel import TravelEntry
from app.models.user import User
from app.services.report_export_service import safe_csv_cell


PersonType = Literal["therapist", "doctor"]
ReportFormat = Literal["csv", "xlsx", "pdf"]

EXPORT_ROW_LIMIT = 25_000
PDF_ROW_LIMIT = 3_000

PERSON_TYPE_PLURAL = {"therapist": "Therapists", "doctor": "Doctors"}
NO_RECORDS_MESSAGE = "No travel expense records found for the selected period."

DETAIL_COLUMNS = (
    "Date",
    "Patient Name",
    "From Address",
    "To Address",
    "KM",
    "Fare",
    "Daily Allowance",
    "Others",
    "Total",
)
ALL_COLUMNS = ("Therapist/Doctor Name", *DETAIL_COLUMNS)


@dataclass
class TravelExpenseRow:
    date: date
    patient_name: str
    from_address: str
    to_address: str
    km: float
    fare: float
    daily_allowance: float
    others: float
    total: float


@dataclass
class TravelExpenseGroup:
    person_id: int
    person_name: str
    rows: list[TravelExpenseRow]

    @property
    def total_km(self) -> float:
        return round(sum(row.km for row in self.rows), 2)

    @property
    def total_fare(self) -> float:
        return round(sum(row.fare for row in self.rows), 2)

    @property
    def total_daily_allowance(self) -> float:
        return round(sum(row.daily_allowance for row in self.rows), 2)

    @property
    def total_others(self) -> float:
        return round(sum(row.others for row in self.rows), 2)

    @property
    def grand_total(self) -> float:
        return round(sum(row.total for row in self.rows), 2)


def resolve_report_period(
    month: str | None,
    start_date: date | None,
    end_date: date | None,
) -> tuple[date, date]:
    """Single source of truth for month/custom-range date math (leap years included)."""
    if month is not None:
        if start_date is not None or end_date is not None:
            raise HTTPException(
                status_code=422,
                detail="Provide either a month or a custom date range, not both.",
            )
        match = re.fullmatch(r"(\d{4})-(\d{2})", month)
        if not match:
            raise HTTPException(
                status_code=422,
                detail="Month must be in YYYY-MM format.",
            )
        year, month_num = int(match.group(1)), int(match.group(2))
        if not 1 <= month_num <= 12:
            raise HTTPException(
                status_code=422,
                detail="Month must be between 01 and 12.",
            )
        last_day = calendar.monthrange(year, month_num)[1]
        return date(year, month_num, 1), date(year, month_num, last_day)

    if start_date is None or end_date is None:
        raise HTTPException(
            status_code=422,
            detail="Provide a month, or both start_date and end_date.",
        )
    if start_date > end_date:
        raise HTTPException(
            status_code=422,
            detail="Start date cannot be after end date.",
        )
    return start_date, end_date


def _period_phrase(start: date, end: date) -> str:
    if start.year == end.year and start.month == end.month:
        month_name = start.strftime("%B").upper()
        return f"{month_name} FROM {start.day} TO {end.day}"
    return f"FROM {start:%d %b %Y} TO {end:%d %b %Y}"


def build_heading(
    person_type: PersonType,
    scope: Literal["individual", "all"],
    person_name: str | None,
    start_date: date,
    end_date: date,
) -> str:
    phrase = _period_phrase(start_date, end_date)
    if scope == "all":
        subject = f"ALL {PERSON_TYPE_PLURAL[person_type].upper()}"
    else:
        subject = f"{(person_name or '').strip().upper()} {person_type.upper()}"
    return f"{subject} TRAVEL EXPENSE {phrase}"


def _sanitize_filename_part(text: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return cleaned or "Report"


def build_filename(
    person_type: PersonType,
    scope: Literal["individual", "all"],
    person_name: str | None,
    start_date: date,
    end_date: date,
    export_format: ReportFormat,
) -> str:
    if scope == "all":
        subject = f"All_{PERSON_TYPE_PLURAL[person_type]}"
    else:
        subject = _sanitize_filename_part(person_name or person_type)
    person_type_label = person_type.title()
    if start_date.year == end_date.year and start_date.month == end_date.month:
        range_part = (
            f"{start_date.strftime('%B')}_{start_date.day:02d}"
            f"-{end_date.day:02d}_{start_date.year}"
        )
    else:
        range_part = f"{start_date:%Y-%m-%d}_to_{end_date:%Y-%m-%d}"
    extension = "xlsx" if export_format == "xlsx" else export_format
    return f"{subject}_{person_type_label}_Travel_Expense_{range_part}.{extension}"


def _therapist_groups(
    db: Session,
    *,
    person_id: int | None,
    start_date: date,
    end_date: date,
) -> list[TravelExpenseGroup]:
    query = (
        db.query(TravelEntry, User.username)
        .join(User, User.id == TravelEntry.therapist_id)
        .filter(func.date(TravelEntry.travel_date) >= start_date)
        .filter(func.date(TravelEntry.travel_date) <= end_date)
    )
    if person_id is not None:
        query = query.filter(TravelEntry.therapist_id == person_id)
    travels = query.order_by(
        TravelEntry.therapist_id,
        TravelEntry.travel_date,
        TravelEntry.id,
    ).all()

    claim_query = db.query(Claim).filter(
        Claim.claim_date >= start_date,
        Claim.claim_date <= end_date,
    )
    if person_id is not None:
        claim_query = claim_query.filter(Claim.therapist_id == person_id)
    daily_allowance_by_day = {
        (claim.therapist_id, claim.claim_date): float(claim.daily_allowance or 0)
        for claim in claim_query.all()
    }

    groups: dict[int, TravelExpenseGroup] = {}
    allowance_assigned_days: set[tuple[int, date]] = set()
    for travel, staff_name in travels:
        travel_day = (
            travel.travel_date.date()
            if isinstance(travel.travel_date, datetime)
            else travel.travel_date
        )
        day_key = (travel.therapist_id, travel_day)
        if day_key in allowance_assigned_days:
            daily_allowance = 0.0
        else:
            daily_allowance = daily_allowance_by_day.get(day_key, 0.0)
            allowance_assigned_days.add(day_key)

        fare = float(travel.travel_fare or 0)
        row = TravelExpenseRow(
            date=travel_day,
            patient_name=travel.patient_name or "N/A",
            from_address=travel.from_address or "",
            to_address=travel.to_address or "",
            km=round(float(travel.total_km or 0), 2),
            fare=round(fare, 2),
            daily_allowance=round(daily_allowance, 2),
            others=0.0,
            total=round(fare + daily_allowance, 2),
        )
        group = groups.setdefault(
            travel.therapist_id,
            TravelExpenseGroup(
                person_id=travel.therapist_id,
                person_name=staff_name,
                rows=[],
            ),
        )
        group.rows.append(row)

    return sorted(groups.values(), key=lambda group: group.person_name.lower())


def _doctor_groups(
    db: Session,
    *,
    person_id: int | None,
    start_date: date,
    end_date: date,
) -> list[TravelExpenseGroup]:
    query = (
        db.query(DoctorExpense, Doctor.name, DoctorVisit.patient_name)
        .join(Doctor, Doctor.id == DoctorExpense.doctor_id)
        .outerjoin(DoctorVisit, DoctorVisit.id == DoctorExpense.visit_id)
        .filter(DoctorExpense.expense_date >= start_date)
        .filter(DoctorExpense.expense_date <= end_date)
    )
    if person_id is not None:
        query = query.filter(DoctorExpense.doctor_id == person_id)
    expenses = query.order_by(
        DoctorExpense.doctor_id,
        DoctorExpense.expense_date,
        DoctorExpense.id,
    ).all()

    groups: dict[int, TravelExpenseGroup] = {}
    for expense, staff_name, patient_name in expenses:
        fare = float(
            expense.approved_amount
            if expense.approved_amount is not None
            else expense.fare or 0
        )
        row = TravelExpenseRow(
            date=expense.expense_date,
            patient_name=patient_name or "N/A",
            from_address=expense.from_location or "",
            to_address=expense.to_location or "",
            km=round(float(expense.distance_km or 0), 2),
            fare=round(fare, 2),
            daily_allowance=0.0,
            others=0.0,
            total=round(fare, 2),
        )
        group = groups.setdefault(
            expense.doctor_id,
            TravelExpenseGroup(
                person_id=expense.doctor_id,
                person_name=staff_name,
                rows=[],
            ),
        )
        group.rows.append(row)

    return sorted(groups.values(), key=lambda group: group.person_name.lower())


def get_travel_expense_groups(
    db: Session,
    *,
    person_type: PersonType,
    person_id: int | None,
    start_date: date,
    end_date: date,
) -> list[TravelExpenseGroup]:
    """Single shared query used by preview, PDF, Excel, and CSV alike."""
    if person_type == "therapist":
        return _therapist_groups(
            db,
            person_id=person_id,
            start_date=start_date,
            end_date=end_date,
        )
    return _doctor_groups(
        db,
        person_id=person_id,
        start_date=start_date,
        end_date=end_date,
    )


def _period_label(start_date: date, end_date: date) -> str:
    return f"{start_date:%d %b %Y} - {end_date:%d %b %Y}"


def build_travel_expense_report(
    db: Session,
    *,
    person_type: PersonType,
    scope: Literal["individual", "all"],
    person_id: int | None,
    person_name: str | None,
    start_date: date,
    end_date: date,
) -> dict:
    groups = get_travel_expense_groups(
        db,
        person_type=person_type,
        person_id=person_id,
        start_date=start_date,
        end_date=end_date,
    )
    row_count = sum(len(group.rows) for group in groups)
    if row_count > EXPORT_ROW_LIMIT:
        raise HTTPException(
            status_code=413,
            detail=(
                f"This report exceeds {EXPORT_ROW_LIMIT:,} rows. "
                "Use a smaller date range."
            ),
        )
    warnings = [] if row_count else [NO_RECORDS_MESSAGE]
    heading = build_heading(person_type, scope, person_name, start_date, end_date)
    return {
        "heading": heading,
        "person_type": person_type,
        "scope": scope,
        "person_id": person_id,
        "person_name": person_name,
        "period_label": _period_label(start_date, end_date),
        "start_date": start_date,
        "end_date": end_date,
        "groups": groups,
        "total_km": round(sum(group.total_km for group in groups), 2),
        "total_fare": round(sum(group.total_fare for group in groups), 2),
        "total_daily_allowance": round(
            sum(group.total_daily_allowance for group in groups), 2
        ),
        "total_others": round(sum(group.total_others for group in groups), 2),
        "grand_total": round(sum(group.grand_total for group in groups), 2),
        "row_count": row_count,
        "generated_at": datetime.now(timezone.utc),
        "warnings": warnings,
    }


def _flat_rows(
    groups: list[TravelExpenseGroup],
    *,
    include_name_column: bool,
) -> list[list]:
    rows: list[list] = []
    for group in groups:
        for row in group.rows:
            base = [
                row.date.isoformat(),
                row.patient_name,
                row.from_address,
                row.to_address,
                row.km,
                row.fare,
                row.daily_allowance,
                row.others,
                row.total,
            ]
            rows.append([group.person_name, *base] if include_name_column else base)
    return rows


def _build_csv(report: dict) -> bytes:
    include_name = report["scope"] == "all"
    headers = ALL_COLUMNS if include_name else DETAIL_COLUMNS
    output = StringIO(newline="")
    writer = csv_module.writer(output)
    writer.writerow(headers)
    if report["row_count"] == 0:
        writer.writerow([NO_RECORDS_MESSAGE] + [""] * (len(headers) - 1))
    else:
        for row in _flat_rows(report["groups"], include_name_column=include_name):
            writer.writerow([safe_csv_cell(value) for value in row])
    return ("﻿" + output.getvalue()).encode("utf-8")


def _build_xlsx(report: dict) -> bytes:
    include_name = report["scope"] == "all"
    headers = ALL_COLUMNS if include_name else DETAIL_COLUMNS
    # 1-indexed openpyxl column positions, offset by 1 when the leading
    # "Therapist/Doctor Name" column is present.
    offset = 1 if include_name else 0
    km_column = 5 + offset
    currency_columns = (6 + offset, 7 + offset, 8 + offset, 9 + offset)
    address_columns = (3 + offset, 4 + offset)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Travel Expense"

    sheet.append([report["heading"]])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(headers)
    )
    sheet.append([f"Period: {report['period_label']}"])
    sheet.append([f"Generated: {report['generated_at'].isoformat()} UTC"])
    sheet.append([])

    header_row_index = sheet.max_row + 1
    sheet.append(list(headers))
    header_fill = PatternFill("solid", fgColor="1B5E20")
    for cell in sheet[header_row_index]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    if report["row_count"] == 0:
        sheet.append([NO_RECORDS_MESSAGE])
        sheet.merge_cells(
            start_row=sheet.max_row,
            start_column=1,
            end_row=sheet.max_row,
            end_column=len(headers),
        )
    else:
        for row in _flat_rows(report["groups"], include_name_column=include_name):
            sheet.append(row)
        blank_cells = ["", "", "", ""] if include_name else ["", "", ""]
        total_row = [
            "GRAND TOTAL",
            *blank_cells,
            report["total_km"],
            report["total_fare"],
            report["total_daily_allowance"],
            report["total_others"],
            report["grand_total"],
        ]
        sheet.append(total_row)
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)

    sheet.freeze_panes = sheet.cell(row=header_row_index + 1, column=1)
    sheet.auto_filter.ref = (
        f"A{header_row_index}:{get_column_letter(len(headers))}{header_row_index}"
    )
    for column_index, header in enumerate(headers, start=1):
        width = 34 if column_index in address_columns else min(
            max(len(header) + 2, 12), 30
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row_index in range(header_row_index + 1, sheet.max_row + 1):
        for column_index in address_columns:
            sheet.cell(row=row_index, column=column_index).alignment = wrap_alignment
        for column_index in currency_columns:
            sheet.cell(row=row_index, column=column_index).number_format = "#,##0.00"
        sheet.cell(row=row_index, column=km_column).number_format = "0.00"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class _NumberedCanvas(pdf_canvas.Canvas):
    """Adds 'Page X of Y' footers, standard reportlab recipe."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_footer(total_pages)
            super().showPage()
        super().save()

    def _draw_footer(self, total_pages: int) -> None:
        self.setFont("Helvetica", 7)
        self.setFillColor(colors.HexColor("#64748B"))
        page_width = self._pagesize[0]
        self.drawRightString(
            page_width - 12 * mm,
            8 * mm,
            f"Page {self._pageNumber} of {total_pages}",
        )
        self.drawString(
            12 * mm,
            8 * mm,
            f"Generated {datetime.now(timezone.utc):%d %b %Y %H:%M} UTC",
        )


def _build_pdf(report: dict) -> bytes:
    if report["row_count"] > PDF_ROW_LIMIT:
        raise HTTPException(
            status_code=413,
            detail=(
                f"PDF export supports up to {PDF_ROW_LIMIT:,} rows. "
                "Use a smaller date range or download XLSX/CSV."
            ),
        )
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=16 * mm,
        title=report["heading"],
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "cell", parent=styles["BodyText"], fontSize=7.5, leading=9
    )
    subtotal_style = ParagraphStyle(
        "subtotal", parent=styles["BodyText"], fontSize=8, leading=10, spaceBefore=2
    )

    story = [
        Paragraph(escape(report["heading"]), styles["Title"]),
        Paragraph(
            f"Period: {escape(report['period_label'])}", styles["BodyText"]
        ),
        Paragraph(
            f"Generated: {report['generated_at']:%d %b %Y %H:%M} UTC",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
    ]

    if report["row_count"] == 0:
        story.append(Paragraph(escape(NO_RECORDS_MESSAGE), styles["BodyText"]))
        document.build(story, canvasmaker=_NumberedCanvas)
        return output.getvalue()

    show_grouped_headers = report["scope"] == "all"
    col_widths_mm = (18, 34, 46, 46, 14, 18, 22, 16, 18)

    for group in report["groups"]:
        if show_grouped_headers:
            story.append(
                Paragraph(escape(group.person_name.upper()), styles["Heading3"])
            )
        table_data = [list(DETAIL_COLUMNS)]
        for row in group.rows:
            table_data.append(
                [
                    row.date.strftime("%d-%m-%Y"),
                    Paragraph(escape(row.patient_name), cell_style),
                    Paragraph(escape(row.from_address), cell_style),
                    Paragraph(escape(row.to_address), cell_style),
                    f"{row.km:.2f} km",
                    f"{row.fare:,.2f}",
                    f"{row.daily_allowance:,.2f}",
                    f"{row.others:,.2f}",
                    f"{row.total:,.2f}",
                ]
            )
        table = Table(
            table_data,
            colWidths=[width * mm for width in col_widths_mm],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E20")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("FONTSIZE", (0, 1), (-1, -1), 7.5),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F8FAFC")],
                    ),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)
        if show_grouped_headers:
            story.append(
                Paragraph(
                    f"<b>{escape(group.person_name.upper())} TOTAL</b> &nbsp;&nbsp; "
                    f"Total KM: {group.total_km:.2f} km &nbsp; "
                    f"Total Fare: {group.total_fare:,.2f} &nbsp; "
                    f"Total Daily Allowance: {group.total_daily_allowance:,.2f} &nbsp; "
                    f"Total Others: {group.total_others:,.2f} &nbsp; "
                    f"Grand Total: {group.grand_total:,.2f}",
                    subtotal_style,
                )
            )
        story.append(Spacer(1, 5 * mm))

    story.append(
        Paragraph(
            f"<b>OVERALL TOTAL</b> &nbsp;&nbsp; "
            f"Total KM: {report['total_km']:.2f} km &nbsp; "
            f"Total Fare: {report['total_fare']:,.2f} &nbsp; "
            f"Total Daily Allowance: {report['total_daily_allowance']:,.2f} &nbsp; "
            f"Total Others: {report['total_others']:,.2f} &nbsp; "
            f"Grand Total: {report['grand_total']:,.2f}",
            styles["Heading4"],
        )
    )

    document.build(story, canvasmaker=_NumberedCanvas)
    return output.getvalue()


def build_travel_expense_export_response(
    report: dict,
    export_format: ReportFormat,
) -> Response:
    if export_format == "xlsx":
        content = _build_xlsx(report)
        media_type = (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    elif export_format == "pdf":
        content = _build_pdf(report)
        media_type = "application/pdf"
    else:
        content = _build_csv(report)
        media_type = "text/csv; charset=utf-8"

    filename = build_filename(
        report["person_type"],
        report["scope"],
        report["person_name"],
        report["start_date"],
        report["end_date"],
        export_format,
    )
    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Report-Row-Count": str(report["row_count"]),
            "Cache-Control": "private, no-store",
            "X-Content-Type-Options": "nosniff",
        },
    )
