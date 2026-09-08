import unittest
from datetime import date, datetime, timedelta
from io import BytesIO

from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from openpyxl import load_workbook

from app.database import Base, get_db
from app.models.claim import Claim
from app.models.doctor import Doctor
from app.models.doctor_claim import DoctorClaim
from app.models.travel import TravelEntry
from app.models.treatment_schedule import TreatmentSchedule
from app.models.user import User
from app.routers.admin_reports import router
from app.utils.auth import get_current_user


class AdminReportsApiTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.session_factory = sessionmaker(bind=self.engine)
        self.db = self.session_factory()

        self.admin = User(
            username="Admin",
            email="admin@example.com",
            password_hash="unused",
            role="admin",
            is_active=True,
        )
        self.therapist = User(
            username="Therapist One",
            email="therapist@example.com",
            password_hash="unused",
            role="therapist",
            is_active=True,
        )
        self.other_therapist = User(
            username="Therapist Two",
            email="other@example.com",
            password_hash="unused",
            role="therapist",
            is_active=True,
        )
        doctor_user = User(
            username="Doctor",
            email="doctor@example.com",
            password_hash="unused",
            role="doctor",
            is_active=True,
        )
        self.db.add_all(
            [
                self.admin,
                self.therapist,
                self.other_therapist,
                doctor_user,
            ]
        )
        self.db.flush()

        doctor = Doctor(user_id=doctor_user.id, name="Doctor")
        self.doctor = doctor
        self.db.add(doctor)
        self.db.flush()

        today = date.today()
        completed = TreatmentSchedule(
            patient_name="Private",
            doctor_id=doctor.id,
            therapist_id=self.therapist.id,
            treatment_name="Treatment",
            patient_address="Private",
            schedule_type="one_time",
            treatment_date=today - timedelta(days=1),
            status="completed",
            completed_at=datetime.combine(
                today - timedelta(days=1),
                datetime.min.time(),
            ),
        )
        scheduled = TreatmentSchedule(
            patient_name="Private",
            doctor_id=doctor.id,
            therapist_id=self.therapist.id,
            treatment_name="Treatment",
            patient_address="Private",
            schedule_type="one_time",
            treatment_date=today,
            status="scheduled",
        )
        self.db.add_all([completed, scheduled])
        self.db.flush()

        self.db.add_all(
            [
                TravelEntry(
                    therapist_id=self.therapist.id,
                    schedule_id=completed.id,
                    travel_date=datetime.combine(
                        today - timedelta(days=1),
                        datetime.min.time(),
                    ),
                    from_address="Private",
                    to_address="Private",
                    total_km=12.5,
                    per_km_rate=8,
                    travel_fare=100,
                    patient_visited=True,
                    status="submitted",
                ),
                Claim(
                    therapist_id=self.therapist.id,
                    claim_date=today - timedelta(days=1),
                    total_km=12.5,
                    travel_total=100,
                    daily_allowance=150,
                    grand_total=250,
                    status="pending",
                ),
                Claim(
                    therapist_id=self.other_therapist.id,
                    claim_date=today - timedelta(days=2),
                    total_km=5,
                    travel_total=40,
                    grand_total=40,
                    status="approved",
                ),
            ]
        )
        self.db.commit()

        self.current_user = self.admin
        app = FastAPI()
        app.include_router(router)
        app.dependency_overrides[get_db] = lambda: self.db
        app.dependency_overrides[get_current_user] = (
            lambda: self.current_user
        )
        self.client = TestClient(app)

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def test_returns_aggregated_report_without_patient_details(self):
        response = self.client.get("/admin-reports/overview")

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["kpis"]["todays_treatments"], 1)
        self.assertEqual(body["kpis"]["completed_treatments"], 1)
        self.assertEqual(body["kpis"]["total_claims"], 2)
        self.assertEqual(body["kpis"]["pending_claims"], 1)
        self.assertEqual(body["kpis"]["total_km"], 12.5)
        self.assertEqual(body["kpis"]["total_travel_amount"], 100)
        self.assertEqual(
            body["top_therapists"][0]["therapist_name"],
            "Therapist One",
        )
        self.assertNotIn("patient_name", str(body))

    def test_applies_therapist_and_claim_status_filters(self):
        response = self.client.get(
            "/admin-reports/overview",
            params={
                "therapist_id": self.therapist.id,
                "status": "pending",
            },
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["kpis"]["total_claims"], 1)
        self.assertEqual(body["kpis"]["approved_claims"], 0)
        self.assertEqual(len(body["top_therapists"]), 1)

    def test_overview_claim_kpis_include_doctors_and_match_register_scope(self):
        self.db.add(
            DoctorClaim(
                doctor_id=self.doctor.id,
                claim_date=date.today(),
                total_amount=500,
                expense_count=2,
                status="pending",
                submitted_at=datetime.now(),
            )
        )
        self.db.commit()

        overview = self.client.get("/admin-reports/overview")
        export = self.client.get("/admin-reports/claims/export")

        self.assertEqual(overview.status_code, 200)
        self.assertEqual(export.status_code, 200)
        self.assertEqual(overview.json()["kpis"]["total_claims"], 3)
        self.assertEqual(overview.json()["kpis"]["pending_claims"], 2)
        self.assertEqual(
            overview.json()["kpis"]["total_claims"],
            int(export.headers["x-report-row-count"]),
        )
        highest = next(
            insight
            for insight in overview.json()["insights"]
            if insight["key"] == "highest-claim"
        )
        self.assertEqual(highest["value"], "INR 500.00")

    def test_rejects_invalid_date_range(self):
        response = self.client.get(
            "/admin-reports/overview",
            params={
                "from_date": "2026-02-02",
                "to_date": "2026-02-01",
            },
        )

        self.assertEqual(response.status_code, 422)

    def test_claim_register_export_contains_therapist_and_doctor_rows(self):
        self.db.add(
            DoctorClaim(
                doctor_id=self.doctor.id,
                claim_date=date.today(),
                total_amount=500,
                expense_count=2,
                status="pending",
                submitted_at=datetime.now(),
            )
        )
        self.db.commit()

        response = self.client.get("/admin-reports/claims/export")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.headers["content-type"].startswith("text/csv"))
        self.assertEqual(response.headers["x-report-row-count"], "3")
        self.assertIn("claim-register-", response.headers["content-disposition"])
        self.assertIn("Therapist,Therapist One", response.text)
        self.assertIn("Doctor,Doctor", response.text)
        self.assertNotIn("Private", response.text)

    def test_claim_register_export_applies_role_and_status(self):
        response = self.client.get(
            "/admin-reports/claims/export",
            params={"role": "therapist", "status": "approved"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["x-report-row-count"], "1")
        self.assertIn("Therapist,Therapist Two", response.text)
        self.assertNotIn("Therapist One", response.text)

    def test_claim_register_xlsx_uses_same_snapshot_rows(self):
        response = self.client.get(
            "/admin-reports/claims/export",
            params={"format": "xlsx", "role": "therapist"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.headers["x-report-row-count"], "2")
        self.assertTrue(
            response.headers["content-type"].startswith(
                "application/vnd.openxmlformats-officedocument"
            )
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Summary", "Claims"])
        claims = workbook["Claims"]
        self.assertEqual(claims.max_row, 3)
        self.assertEqual(claims["A2"].value, "Therapist")
        self.assertNotIn("Private", str(list(claims.values)))

    def test_claim_register_pdf_is_server_generated_and_privacy_safe(self):
        response = self.client.get(
            "/admin-reports/claims/export",
            params={"format": "pdf", "status": "pending"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.headers["content-type"], "application/pdf")
        self.assertEqual(response.headers["x-report-row-count"], "1")
        self.assertTrue(response.content.startswith(b"%PDF-"))
        self.assertNotIn(b"Private", response.content)

    def test_claim_register_rejects_unsupported_format(self):
        response = self.client.get(
            "/admin-reports/claims/export",
            params={"format": "xls"},
        )

        self.assertEqual(response.status_code, 422)

    def test_non_admin_cannot_view_reports(self):
        self.current_user = self.therapist

        response = self.client.get("/admin-reports/overview")

        self.assertEqual(response.status_code, 403)

        export_response = self.client.get("/admin-reports/claims/export")
        self.assertEqual(export_response.status_code, 403)


if __name__ == "__main__":
    unittest.main()
